#!/usr/bin/env python3
"""Generate the Legal Mind opruimverantwoordelijkheid Excel file."""
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def generate(schedule: dict, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Weekindeling"

    ORANGE = "E86832"
    DARK_BG = "2B2B2B"
    WARM_BG = "F5F1EE"
    LIGHT_CARD = "EDEDED"
    WHITE = "FFFFFF"
    SUBTLE_GREY = "999999"

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 52

    thin_border = Border(
        left=Side(style="thin", color=LIGHT_CARD),
        right=Side(style="thin", color=LIGHT_CARD),
        top=Side(style="thin", color=LIGHT_CARD),
        bottom=Side(style="thin", color=LIGHT_CARD),
    )

    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value = "Legal Mind — Opruimverantwoordelijkheid"
    c.font = Font(name="Calibri", bold=True, color=DARK_BG, size=16)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:C2")
    c = ws["A2"]
    c.value = "Wekelijks herhalend  ·  Koffiecorner opgeruimd houden (door de dag heen) & vaat"
    c.font = Font(name="Calibri", color=SUBTLE_GREY, size=10)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 22

    for col in range(1, 4):
        cell = ws.cell(row=3, column=col)
        cell.fill = PatternFill("solid", fgColor=ORANGE)
    ws.row_dimensions[3].height = 4

    ws.row_dimensions[4].height = 8

    headers = ["Dag", "Verantwoordelijke(n)", "Taak"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col)
        cell.value = h
        cell.font = Font(name="Calibri", bold=True, color=WHITE, size=11)
        cell.fill = PatternFill("solid", fgColor=DARK_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[5].height = 30

    row = 6
    alt = False
    for dag, namen in schedule.items():
        fill = PatternFill("solid", fgColor=LIGHT_CARD) if alt else PatternFill("solid", fgColor=WHITE)
        alt = not alt

        cell = ws.cell(row=row, column=1)
        cell.value = dag
        cell.font = Font(name="Calibri", bold=True, color=ORANGE, size=11)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

        cell = ws.cell(row=row, column=2)
        cell.value = " & ".join(namen) if isinstance(namen, list) else namen
        cell.font = Font(name="Calibri", bold=True, color=DARK_BG, size=11)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

        cell = ws.cell(row=row, column=3)
        cell.value = "Koffiecorner opgeruimd houden (door de dag heen) + vaat"
        cell.font = Font(name="Calibri", color=DARK_BG, size=11)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border

        ws.row_dimensions[row].height = 28
        row += 1

    row += 1
    ws.merge_cells(f"A{row}:C{row}")
    c = ws.cell(row=row, column=1)
    c.value = "Bij afwezigheid onderling ruilen en even doorgeven."
    c.font = Font(name="Calibri", italic=True, size=10, color=SUBTLE_GREY)

    wb.save(output_path)
    print(f"Rooster opgeslagen: {output_path}")

if __name__ == "__main__":
    schedule = json.loads(sys.argv[1])
    output_path = sys.argv[2] if len(sys.argv) > 2 else "opruimverantwoordelijkheid.xlsx"
    generate(schedule, output_path)
