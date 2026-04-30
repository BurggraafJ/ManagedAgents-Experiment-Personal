"""
Licentie-analyse verwerker voor Legal Mind — v4.

Genereert een leesbaar Excel-rapport in 5 tabbladen vanuit één of meerdere
AFAS-exports. Werkt met categorieën en buckets gericht op intern vs.
extern (klant) gebruik.

Categorieën (intern in leveranciers.json):
  - saas_licentie   → Software bucket
  - abonnement      → Software bucket
  - ai_licentie     → AI-tools bucket
  - tokens_intern   → Tokens Intern bucket (engineering/dienstverbetering team)
  - tokens_extern   → Tokens Extern bucket (productie/cloud klant — Moonlit, Azure, GCP, Anthropic API, AWS)

Status per leverancier: 'actief' of 'stopgezet'. Stopgezet telt NIET mee
in maandgemiddeldes en niet in actieve totalen — wel apart zichtbaar.

Tabbladen:
  1. Samenvatting       — KPI's + per maand + top 10 + reeds-stopgezet info
  2. AI & Tokens Intern — interne AI/tokens + per team
  3. Cloud / Tokens Extern — productiekosten richting klant
  4. Beslislijst        — alleen verplaatsingsoverzicht (staat hier, hoort hier)
  5. Alle mutaties      — ruwe data verrijkt + legenda

Multi-source aanroep:
  python verwerk_licenties.py "abon.xlsx=Abonnementen,cloud.xlsx=Inkoop cloud" \\
      leveranciers.json teams.json output.xlsx

Single-source blijft werken (krijgt label 'Hoofdgrootboek').
"""

from __future__ import annotations

import json
import sys
from datetime import date as _date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Constanten
# ---------------------------------------------------------------------------

MAAND_NAMEN = {
    1: "Januari", 2: "Februari", 3: "Maart", 4: "April",
    5: "Mei", 6: "Juni", 7: "Juli", 8: "Augustus",
    9: "September", 10: "Oktober", 11: "November", 12: "December",
}

# Interne codes -> buckets in de UI (5 buckets nu)
BUCKET_VAN_CODE = {
    "saas_licentie": "Software",
    "abonnement": "Software",
    "ai_licentie": "AI-tools",
    "tokens_intern": "Tokens Intern",
    "tokens_extern": "Tokens Extern",
    "onbekend": "Te classificeren",
}

# Volgorde voor weergave
BUCKETS_VOLGORDE = ["Software", "AI-tools", "Tokens Intern", "Tokens Extern"]

# Welke buckets gaan onder welk hoofd-tabblad
INTERN_BUCKETS = ["Software", "AI-tools", "Tokens Intern"]
EXTERN_BUCKETS = ["Tokens Extern"]

# Welke grootboekrekening hoort welke bucket?
# Tokens Extern hoort op cloud-grootboek; alle andere op abon-grootboek.
JUISTE_BRON_VAN_BUCKET = {
    "Software": "Abonnementen",
    "AI-tools": "Abonnementen",
    "Tokens Intern": "Abonnementen",
    "Tokens Extern": "Inkoop cloud",
}

# Styling
ARIAL = "Arial"
HEADER_FONT = Font(name=ARIAL, bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
DATA_FONT = Font(name=ARIAL, size=10)
BOLD_FONT = Font(name=ARIAL, bold=True, size=10)
ITALIC_FONT = Font(name=ARIAL, italic=True, size=9, color="666666")

KPI_LABEL_FONT = Font(name=ARIAL, bold=True, color="FFFFFF", size=10)
KPI_LABEL_FILL = PatternFill("solid", fgColor="1F4E79")
KPI_VALUE_FONT = Font(name=ARIAL, bold=True, size=18, color="1F4E79")
KPI_VALUE_FILL = PatternFill("solid", fgColor="EAF2FA")

SECTION_TITLE_FONT = Font(name=ARIAL, bold=True, color="1F4E79", size=12)
SECTION_FILL = PatternFill("solid", fgColor="EAF2FA")
TOTAL_FILL = PatternFill("solid", fgColor="D9E2F3")
GROEN_FILL = PatternFill("solid", fgColor="E2EFDA")
GEEL_FILL = PatternFill("solid", fgColor="FFF2CC")
ORANJE_FILL = PatternFill("solid", fgColor="FCE4D6")
GRIJS_FILL = PatternFill("solid", fgColor="F2F2F2")
STOPGEZET_FILL = PatternFill("solid", fgColor="EDEDED")

BUCKET_FILL = {
    "Software": None,
    "AI-tools": PatternFill("solid", fgColor="E8D5F5"),
    "Tokens Intern": PatternFill("solid", fgColor="FFE699"),
    "Tokens Extern": PatternFill("solid", fgColor="FFC000"),
    "Te classificeren": PatternFill("solid", fgColor="F2F2F2"),
}

THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)

EURO_HEEL = "€ #,##0"
EURO_CENT = "€ #,##0.00"
PCT_FMT = "0.0%"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def match_leverancier(boeking_raw, leveranciers):
    if pd.isna(boeking_raw):
        return None
    boeking_lower = str(boeking_raw).strip().lower()
    candidates = []
    for lev in leveranciers:
        for patroon in lev["patronen"]:
            if patroon.lower() in boeking_lower:
                candidates.append((len(patroon), lev))
    if not candidates:
        return None
    candidates.sort(key=lambda x: -x[0])
    return candidates[0][1]


def trend_indicator(per_period):
    keys = sorted(per_period.keys())
    if len(keys) < 2:
        return "—"
    first, last = per_period[keys[0]], per_period[keys[-1]]
    if first == 0:
        return "↑" if last > 0 else "—"
    ratio = last / first
    if ratio > 1.1:
        return "↑"
    if ratio < 0.9:
        return "↓"
    return "→"


def auto_width(ws, max_width=42, min_width=10):
    for col_idx in range(1, ws.max_column + 1):
        max_len = min_width
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 80),
                                min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    if isinstance(cell.value, float):
                        rendered = f"{cell.value:,.0f}"
                    else:
                        rendered = str(cell.value)
                    if len(rendered) > max_len:
                        max_len = len(rendered)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, max_width)


def style_header_row(ws, row_num):
    for cell in ws[row_num]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def write_section_title(ws, text):
    ws.append([text])
    cell = ws.cell(row=ws.max_row, column=1)
    cell.font = SECTION_TITLE_FONT
    cell.fill = SECTION_FILL


def write_blank(ws):
    ws.append([])


# ---------------------------------------------------------------------------
# Verwerking
# ---------------------------------------------------------------------------

def verwerk(input_spec, lev_path, teams_path, output_path):
    lev_data = load_json(lev_path)
    leveranciers = lev_data["leveranciers"]
    teams_data = load_json(teams_path)

    # Multi-source inlezen
    if "=" in str(input_spec) or "," in str(input_spec):
        bronnen = []
        for stuk in str(input_spec).split(","):
            stuk = stuk.strip()
            if "=" in stuk:
                pad, label = stuk.split("=", 1)
                bronnen.append((pad.strip(), label.strip()))
            else:
                bronnen.append((stuk, "Hoofdgrootboek"))
    else:
        bronnen = [(str(input_spec), "Hoofdgrootboek")]

    delen = []
    for pad, label in bronnen:
        d = pd.read_excel(pad, engine="calamine")
        d["Bron"] = label
        delen.append(d)
    df = pd.concat(delen, ignore_index=True)

    # Match leveranciers
    matches = df["Boeking"].apply(lambda x: match_leverancier(x, leveranciers))
    df["Leverancier"] = matches.apply(lambda x: x["naam"] if x else "Onbekend")
    df["Team"] = matches.apply(lambda x: x["team"] if x else "Onbekend")
    df["Categorie"] = matches.apply(lambda x: x["categorie"] if x else "onbekend")
    df["Bucket"] = df["Categorie"].map(BUCKET_VAN_CODE)
    df["Per_seat"] = matches.apply(lambda x: bool(x.get("per_seat")) if x else False)
    df["Status"] = matches.apply(lambda x: x.get("status", "actief") if x else "actief")
    df["Notitie_lev"] = matches.apply(lambda x: x.get("notities", "") if x else "")
    df["BTW_incl"] = matches.apply(lambda x: bool(x.get("btw_inclusief", False)) if x else False)

    # BTW eraf voor leveranciers waarvan we weten dat het bedrag incl 21% BTW is
    df["Saldo_incl"] = df["Saldo"]  # bewaar origineel
    df["Saldo"] = df.apply(
        lambda r: r["Saldo"] / 1.21 if r["BTW_incl"] else r["Saldo"], axis=1
    )
    df["Maand_nr"] = df["Per."].astype(int)
    df["Maand"] = df["Maand_nr"].map(MAAND_NAMEN).fillna("Correctie")

    # Reguliere maanden = maanden waarin daadwerkelijk geboekt is, op basis
    # van de Boekstuk-datum, alleen tot/met de huidige maand.
    vandaag = _date.today()
    df["Boek_dt"] = pd.to_datetime(df["Boekstuk"])
    df["Boek_maand"] = df["Boek_dt"].dt.month
    df["Boek_jaar"] = df["Boek_dt"].dt.year

    werkelijk = sorted(
        df[
            (df["Boek_jaar"] < vandaag.year)
            | ((df["Boek_jaar"] == vandaag.year)
               & (df["Boek_maand"] <= vandaag.month))
        ]["Boek_maand"].unique()
    )
    aanwezige = sorted(df["Maand_nr"].unique())
    reguliere_maanden = [m for m in werkelijk if 1 <= m <= 12]
    correctie_periodes = [p for p in aanwezige if p not in reguliere_maanden]
    n_maanden = max(len(reguliere_maanden), 1)
    totaal_personen = teams_data["totaal_unieke_personen"]

    df_reg = df[df["Maand_nr"].isin(reguliere_maanden)]
    df_corr = df[df["Maand_nr"].isin(correctie_periodes)]

    # Splitsen op status: actieve mutaties zijn de basis voor maandgemiddeldes
    df_reg_actief = df_reg[df_reg["Status"] == "actief"]
    df_reg_stopgezet = df_reg[df_reg["Status"] == "stopgezet"]

    # Identificeer "verkeerd geboekt": bucket vereist andere bron dan waar mutatie staat
    def _is_verkeerd(row):
        gewenst = JUISTE_BRON_VAN_BUCKET.get(row["Bucket"])
        if gewenst is None:
            return False
        # case-insensitive match op label
        return gewenst.lower() not in str(row["Bron"]).lower()
    df_reg_actief = df_reg_actief.copy()
    df_reg_actief["Verkeerd_geboekt"] = df_reg_actief.apply(_is_verkeerd, axis=1)
    # Voeg deze info ook toe aan df_reg voor consistente verwerking
    df_reg = df_reg.copy()
    df_reg["Verkeerd_geboekt"] = df_reg.apply(_is_verkeerd, axis=1)
    df = df.copy()
    df["Verkeerd_geboekt"] = df.apply(_is_verkeerd, axis=1)

    # Totalen — alleen actief
    totaal_actief_ytd = float(df_reg_actief["Saldo"].sum())
    totaal_stopgezet_ytd = float(df_reg_stopgezet["Saldo"].sum())

    bucket_actief_ytd = {b: 0.0 for b in BUCKETS_VOLGORDE + ["Te classificeren"]}
    for b in df_reg_actief["Bucket"].unique():
        bucket_actief_ytd[b] = float(df_reg_actief[df_reg_actief["Bucket"] == b]["Saldo"].sum())

    intern_ytd = sum(bucket_actief_ytd.get(b, 0) for b in INTERN_BUCKETS)
    extern_ytd = sum(bucket_actief_ytd.get(b, 0) for b in EXTERN_BUCKETS)

    intern_per_maand = intern_ytd / n_maanden
    extern_per_maand = extern_ytd / n_maanden

    wb = Workbook()

    # =====================================================================
    # TAB 1 — Samenvatting
    # =====================================================================
    ws1 = wb.active
    ws1.title = "Samenvatting"
    ws1.sheet_view.showGridLines = False

    ws1.append([f"Licentie-analyse YTD — Legal Mind ({len(reguliere_maanden)} maand{'en' if len(reguliere_maanden) != 1 else ''})"])
    title_cell = ws1.cell(row=1, column=1)
    title_cell.font = Font(name=ARIAL, bold=True, size=14, color="1F4E79")
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    write_blank(ws1)

    # KPI-tegels (4): Totaal actief YTD, Intern/mnd, Extern/mnd, Totaal/mnd
    kpi_label_row = 3
    kpi_value_row = 4
    kpis = [
        ("Totaal actief YTD", totaal_actief_ytd),
        ("Intern / maand", intern_per_maand),
        ("Extern (klant) / maand", extern_per_maand),
        ("Totaal / maand", intern_per_maand + extern_per_maand),
    ]
    for i, (label, value) in enumerate(kpis):
        col = i * 2 + 1
        ws1.merge_cells(start_row=kpi_label_row, start_column=col,
                        end_row=kpi_label_row, end_column=col + 1)
        ws1.merge_cells(start_row=kpi_value_row, start_column=col,
                        end_row=kpi_value_row, end_column=col + 1)
        lc = ws1.cell(row=kpi_label_row, column=col, value=label)
        lc.font = KPI_LABEL_FONT
        lc.fill = KPI_LABEL_FILL
        lc.alignment = Alignment(horizontal="center", vertical="center")
        vc = ws1.cell(row=kpi_value_row, column=col, value=value)
        vc.font = KPI_VALUE_FONT
        vc.fill = KPI_VALUE_FILL
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.number_format = EURO_HEEL
    ws1.row_dimensions[kpi_value_row].height = 38
    # Spring naar rij na de KPI value rij
    while len(ws1["A"]) < kpi_value_row + 1:
        ws1.append([None])
    write_blank(ws1)

    # Per maand totaal
    write_section_title(ws1, "Per maand totaal (alleen actieve leveranciers)")
    header = ["Bucket"] + [MAAND_NAMEN[m] for m in reguliere_maanden] + ["Gem./mnd", "Totaal YTD"]
    ws1.append(header)
    style_header_row(ws1, ws1.max_row)

    for bucket in BUCKETS_VOLGORDE + ["Te classificeren"]:
        b_df = df_reg_actief[df_reg_actief["Bucket"] == bucket]
        if len(b_df) == 0 and bucket == "Te classificeren":
            continue
        row = [bucket]
        for m in reguliere_maanden:
            v = float(b_df[b_df["Maand_nr"] == m]["Saldo"].sum())
            row.append(v)
        ytd = sum(row[1:1 + len(reguliere_maanden)])
        gem = ytd / n_maanden
        row += [gem, ytd]
        ws1.append(row)
        fill = BUCKET_FILL.get(bucket)
        for cell in ws1[ws1.max_row]:
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill
            if isinstance(cell.value, float):
                cell.number_format = EURO_HEEL

    # Totaal-rij
    tot_row = ["TOTAAL"]
    for m in reguliere_maanden:
        tot_row.append(float(df_reg_actief[df_reg_actief["Maand_nr"] == m]["Saldo"].sum()))
    tot_row += [totaal_actief_ytd / n_maanden, totaal_actief_ytd]
    ws1.append(tot_row)
    for cell in ws1[ws1.max_row]:
        cell.font = BOLD_FONT
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
        if isinstance(cell.value, float):
            cell.number_format = EURO_HEEL

    if df_corr.shape[0] > 0:
        correctie_totaal = float(df_corr["Saldo"].sum())
        ws1.append([f"⚠ Correctieboekingen / vooruit-geboekt (periode {','.join(str(p) for p in correctie_periodes)})",
                    "", "", "", "", "", correctie_totaal])
        for cell in ws1[ws1.max_row]:
            cell.font = ITALIC_FONT
            if isinstance(cell.value, float):
                cell.number_format = EURO_HEEL

    if reguliere_maanden and reguliere_maanden[-1] == vandaag.month and \
       df["Boek_jaar"].max() == vandaag.year:
        ws1.append([f"ℹ️ {MAAND_NAMEN[vandaag.month]} loopt nog — totaal van die maand is mogelijk nog niet compleet (cloud-facturen komen vaak laat in de maand binnen)"])
        for cell in ws1[ws1.max_row]:
            cell.font = ITALIC_FONT

    # Detecteer uitschieters per bucket (maand > 1.5x gemiddelde van overige maanden)
    if len(reguliere_maanden) >= 2:
        for bucket in BUCKETS_VOLGORDE:
            b_df = df_reg_actief[df_reg_actief["Bucket"] == bucket]
            if len(b_df) == 0:
                continue
            per_maand = {m: float(b_df[b_df["Maand_nr"] == m]["Saldo"].sum())
                         for m in reguliere_maanden}
            for m, v in per_maand.items():
                # Skip lopende maand (incomplete data)
                if m == vandaag.month and df["Boek_jaar"].max() == vandaag.year:
                    continue
                overige = [per_maand[mm] for mm in reguliere_maanden
                           if mm != m and not (mm == vandaag.month and df["Boek_jaar"].max() == vandaag.year)]
                if not overige:
                    continue
                gem_overige = sum(overige) / len(overige)
                if gem_overige > 0 and v > 1.5 * gem_overige:
                    # Vind top-3 leveranciers in die maand+bucket
                    sub = b_df[b_df["Maand_nr"] == m]
                    top = sub.groupby("Leverancier")["Saldo"].sum().sort_values(ascending=False).head(3)
                    top_str = ", ".join([f"{lev} (€{val:,.0f})" for lev, val in top.items()])
                    ws1.append([f"⚠️ Uitschieter: {bucket} in {MAAND_NAMEN[m]} = € {v:,.0f} "
                                f"(gem. overige maanden: € {gem_overige:,.0f}). "
                                f"Top: {top_str}"])
                    for cell in ws1[ws1.max_row]:
                        cell.font = ITALIC_FONT
                        cell.fill = GEEL_FILL

    write_blank(ws1)
    write_blank(ws1)

    # Top 10 actieve uitgaven YTD
    write_section_title(ws1, "Top 10 actieve uitgaven YTD")
    ws1.append(["#", "Leverancier", "Bucket", "Team", "Totaal YTD", "% van totaal", "Trend"])
    style_header_row(ws1, ws1.max_row)

    lev_totalen = df_reg_actief.groupby(["Leverancier", "Bucket", "Team"])["Saldo"].sum().reset_index()
    lev_totalen = lev_totalen.sort_values("Saldo", ascending=False).head(10)

    for i, (_, r) in enumerate(lev_totalen.iterrows(), 1):
        sub = df_reg_actief[df_reg_actief["Leverancier"] == r["Leverancier"]]
        per_period = sub.groupby("Maand_nr")["Saldo"].sum().to_dict()
        trend = trend_indicator(per_period)
        pct = r["Saldo"] / totaal_actief_ytd if totaal_actief_ytd else 0
        ws1.append([i, r["Leverancier"], r["Bucket"], r["Team"], float(r["Saldo"]), pct, trend])
        fill = BUCKET_FILL.get(r["Bucket"])
        for cell in ws1[ws1.max_row]:
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill
        ws1.cell(row=ws1.max_row, column=5).number_format = EURO_HEEL
        ws1.cell(row=ws1.max_row, column=6).number_format = PCT_FMT
        ws1.cell(row=ws1.max_row, column=7).alignment = Alignment(horizontal="center")

    write_blank(ws1)
    write_blank(ws1)

    # ✅ Reeds stopgezet
    if len(df_reg_stopgezet) > 0:
        write_section_title(ws1, "✅ Reeds stopgezet (informatief — telt niet mee in totalen)")
        ws1.append(["Leverancier", "Bucket", "Totaal YTD (historisch)", "Notitie"])
        style_header_row(ws1, ws1.max_row)
        stop_per_lev = df_reg_stopgezet.groupby(["Leverancier", "Bucket"])["Saldo"].sum().reset_index()
        for _, r in stop_per_lev.sort_values("Saldo", ascending=False).iterrows():
            notitie = ""
            for lev in leveranciers:
                if lev["naam"] == r["Leverancier"]:
                    notitie = lev.get("notities", "")
                    break
            ws1.append([r["Leverancier"], r["Bucket"], float(r["Saldo"]), notitie])
            for cell in ws1[ws1.max_row]:
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                cell.fill = STOPGEZET_FILL
            ws1.cell(row=ws1.max_row, column=3).number_format = EURO_HEEL
        ws1.append(["TOTAAL stopgezet (historisch YTD)", "", totaal_stopgezet_ytd, ""])
        for cell in ws1[ws1.max_row]:
            cell.font = BOLD_FONT
            cell.fill = TOTAL_FILL
            cell.border = THIN_BORDER
        ws1.cell(row=ws1.max_row, column=3).number_format = EURO_HEEL

    write_blank(ws1)
    write_blank(ws1)

    # ❓ Te classificeren
    onbekend_groep = df_reg_actief[df_reg_actief["Leverancier"] == "Onbekend"]
    if len(onbekend_groep) > 0:
        write_section_title(ws1, "❓ Te classificeren — onbekende leveranciers")
        ws1.append(["Boeking origineel", "Bron", "Totaal YTD"])
        style_header_row(ws1, ws1.max_row)
        for boeking, sub in onbekend_groep.groupby("Boeking"):
            tot = float(sub["Saldo"].sum())
            bron_str = ", ".join(sorted(sub["Bron"].unique()))
            ws1.append([str(boeking).strip(), bron_str, tot])
            for cell in ws1[ws1.max_row]:
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                cell.fill = GEEL_FILL
            ws1.cell(row=ws1.max_row, column=3).number_format = EURO_HEEL

    auto_width(ws1, max_width=40)
    for col in range(1, 9):
        cur = ws1.column_dimensions[get_column_letter(col)].width or 12
        ws1.column_dimensions[get_column_letter(col)].width = max(cur, 16)
    ws1.freeze_panes = "A6"

    # =====================================================================
    # TAB 2 — AI & Tokens Intern + Per team
    # =====================================================================
    ws2 = wb.create_sheet("AI & Tokens Intern")
    ws2.sheet_view.showGridLines = False

    ws2.append(["AI & Tokens Intern — wat we als team verbruiken"])
    ws2.cell(row=1, column=1).font = Font(name=ARIAL, bold=True, size=14, color="1F4E79")
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    write_blank(ws2)

    df_intern = df_reg_actief[df_reg_actief["Bucket"].isin(INTERN_BUCKETS)]

    # Sectie A — AI-tools (vaste abonnementen)
    write_section_title(ws2, "A. AI-tools (vaste abonnementen)")
    ws2.append(["Tool", "Team", "Per seat?", "Gem./maand", "Aantal seats",
                "Per persoon/maand", "Trend"])
    style_header_row(ws2, ws2.max_row)

    ai_tools_df = df_intern[df_intern["Bucket"] == "AI-tools"]
    ai_tools_total_pm = 0.0
    for lev_naam in sorted(ai_tools_df["Leverancier"].unique()):
        lev_df = ai_tools_df[ai_tools_df["Leverancier"] == lev_naam]
        team = lev_df["Team"].iloc[0]
        is_per_seat = bool(lev_df["Per_seat"].iloc[0])
        gem_pm = float(lev_df["Saldo"].sum()) / n_maanden
        per_period = lev_df.groupby("Maand_nr")["Saldo"].sum().to_dict()
        trend = trend_indicator(per_period)
        seats = totaal_personen if (is_per_seat and team == "Organisatiebreed") else \
                (teams_data["teams"].get(team, {}).get("aantal_personen") if is_per_seat else None)
        per_persoon = (gem_pm / seats) if (seats and seats > 0) else None
        ws2.append([lev_naam, team, "Ja" if is_per_seat else "Nee", gem_pm,
                    seats if seats else "—",
                    per_persoon if per_persoon is not None else "",
                    trend])
        for cell in ws2[ws2.max_row]:
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.fill = BUCKET_FILL["AI-tools"]
        ws2.cell(row=ws2.max_row, column=4).number_format = EURO_HEEL
        if per_persoon is not None:
            ws2.cell(row=ws2.max_row, column=6).number_format = EURO_HEEL
        ws2.cell(row=ws2.max_row, column=7).alignment = Alignment(horizontal="center")
        ai_tools_total_pm += gem_pm

    ws2.append(["Subtotaal AI-tools", "", "", ai_tools_total_pm, "", "", ""])
    for cell in ws2[ws2.max_row]:
        cell.font = BOLD_FONT
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
    ws2.cell(row=ws2.max_row, column=4).number_format = EURO_HEEL

    write_blank(ws2)

    # Sectie B — Tokens Intern (variabel)
    write_section_title(ws2, "B. Tokens Intern (variabel — engineering/dienstverbetering door team)")
    ws2.append(["Provider", "Team", "Gem./maand", "Totaal YTD", "Trend", "Notitie"])
    style_header_row(ws2, ws2.max_row)

    tokens_intern_df = df_intern[df_intern["Bucket"] == "Tokens Intern"]
    tokens_intern_total_pm = 0.0
    for lev_naam in sorted(tokens_intern_df["Leverancier"].unique()):
        lev_df = tokens_intern_df[tokens_intern_df["Leverancier"] == lev_naam]
        team = lev_df["Team"].iloc[0]
        gem_pm = float(lev_df["Saldo"].sum()) / n_maanden
        ytd = float(lev_df["Saldo"].sum())
        per_period = lev_df.groupby("Maand_nr")["Saldo"].sum().to_dict()
        trend = trend_indicator(per_period)
        notitie = lev_df["Notitie_lev"].iloc[0] or ""
        ws2.append([lev_naam, team, gem_pm, ytd, trend, notitie])
        for cell in ws2[ws2.max_row]:
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.fill = BUCKET_FILL["Tokens Intern"]
        ws2.cell(row=ws2.max_row, column=3).number_format = EURO_HEEL
        ws2.cell(row=ws2.max_row, column=4).number_format = EURO_HEEL
        ws2.cell(row=ws2.max_row, column=5).alignment = Alignment(horizontal="center")
        tokens_intern_total_pm += gem_pm

    ws2.append(["Subtotaal Tokens Intern", "", tokens_intern_total_pm,
                float(tokens_intern_df["Saldo"].sum()), "", ""])
    for cell in ws2[ws2.max_row]:
        cell.font = BOLD_FONT
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
    ws2.cell(row=ws2.max_row, column=3).number_format = EURO_HEEL
    ws2.cell(row=ws2.max_row, column=4).number_format = EURO_HEEL

    write_blank(ws2)

    # Sectie C — Software (overig vast intern)
    write_section_title(ws2, "C. Software (overige vaste licenties team)")
    ws2.append(["Tool", "Team", "Per seat?", "Gem./maand", "Aantal seats",
                "Per persoon/maand", "Trend"])
    style_header_row(ws2, ws2.max_row)

    software_df = df_intern[df_intern["Bucket"] == "Software"]
    # Sorteer op gemiddelde per maand, laag naar hoog
    software_lev_totalen = software_df.groupby("Leverancier")["Saldo"].sum().sort_values(ascending=True)
    software_total_pm = 0.0
    for lev_naam in software_lev_totalen.index:
        lev_df = software_df[software_df["Leverancier"] == lev_naam]
        team = lev_df["Team"].iloc[0]
        is_per_seat = bool(lev_df["Per_seat"].iloc[0])
        gem_pm = float(lev_df["Saldo"].sum()) / n_maanden
        per_period = lev_df.groupby("Maand_nr")["Saldo"].sum().to_dict()
        trend = trend_indicator(per_period)
        seats = totaal_personen if (is_per_seat and team == "Organisatiebreed") else \
                (teams_data["teams"].get(team, {}).get("aantal_personen") if is_per_seat else None)
        per_persoon = (gem_pm / seats) if (seats and seats > 0) else None
        ws2.append([lev_naam, team, "Ja" if is_per_seat else "Nee", gem_pm,
                    seats if seats else "—",
                    per_persoon if per_persoon is not None else "",
                    trend])
        for cell in ws2[ws2.max_row]:
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
        ws2.cell(row=ws2.max_row, column=4).number_format = EURO_HEEL
        if per_persoon is not None:
            ws2.cell(row=ws2.max_row, column=6).number_format = EURO_HEEL
        ws2.cell(row=ws2.max_row, column=7).alignment = Alignment(horizontal="center")
        software_total_pm += gem_pm

    ws2.append(["Subtotaal Software", "", "", software_total_pm, "", "", ""])
    for cell in ws2[ws2.max_row]:
        cell.font = BOLD_FONT
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
    ws2.cell(row=ws2.max_row, column=4).number_format = EURO_HEEL

    write_blank(ws2)
    write_blank(ws2)

    # Sectie D — Per team
    write_section_title(ws2, "D. Per team — gemiddelde kosten per maand (incl. toerekening gedeelde licenties)")
    ws2.append([f"Gedeelde licenties (HubSpot voor Sales+Marketing, Pitch voor Sales+CS) "
                f"worden 50/50 verdeeld over de teams. Organisatiebrede licenties (Atlassian, "
                f"Bitwarden, Loom etc.) worden naar rato van personen toegerekend."])
    ws2.cell(row=ws2.max_row, column=1).font = ITALIC_FONT
    ws2.append(["Team", "Personen", "Eigen Software", "Eigen AI/Tokens",
                "Toegerekend", "Totaal/mnd", "Per persoon/mnd"])
    style_header_row(ws2, ws2.max_row)

    teams_lijst = list(teams_data["teams"].keys())

    # Splits gedeelde teams in {team1: ..., team2: ...} mapping
    GEDEELD_SPLIT = {
        "Sales + Marketing": ["Sales", "Marketing"],
        "Sales + Customer Success": ["Sales", "Customer Success"],
    }

    # Bereken organisatiebrede totaal en alloceer per persoon
    org_df = df_intern[df_intern["Team"] == "Organisatiebreed"]
    org_sw = float(org_df[org_df["Bucket"] == "Software"]["Saldo"].sum()) / n_maanden
    org_ai = float(org_df[org_df["Bucket"] == "AI-tools"]["Saldo"].sum()) / n_maanden
    org_tk = float(org_df[org_df["Bucket"] == "Tokens Intern"]["Saldo"].sum()) / n_maanden
    org_pp = (org_sw + org_ai + org_tk) / totaal_personen if totaal_personen > 0 else 0

    # Bereken gedeelde licenties per gedeeld-team -> per onderliggend team
    gedeeld_per_team = {t: 0.0 for t in teams_lijst}
    for combi_team, sub_teams in GEDEELD_SPLIT.items():
        combi_df = df_intern[df_intern["Team"] == combi_team]
        if len(combi_df) == 0:
            continue
        combi_total_pm = float(combi_df["Saldo"].sum()) / n_maanden
        # 50/50 over de subteams
        per_subteam = combi_total_pm / len(sub_teams)
        for st in sub_teams:
            if st in gedeeld_per_team:
                gedeeld_per_team[st] += per_subteam

    # Nu per team: eigen + toerekening
    grand_eigen_sw = 0.0
    grand_eigen_ai = 0.0
    grand_toegerekend = 0.0
    for team_naam in teams_lijst:
        team_info = teams_data["teams"][team_naam]
        n_pers = team_info["aantal_personen"]
        team_df = df_intern[df_intern["Team"] == team_naam]
        eigen_sw = float(team_df[team_df["Bucket"] == "Software"]["Saldo"].sum()) / n_maanden
        eigen_ai_tk = float(team_df[team_df["Bucket"].isin(["AI-tools", "Tokens Intern"])]["Saldo"].sum()) / n_maanden
        toegerekend = gedeeld_per_team[team_naam] + (org_pp * n_pers)
        tot = eigen_sw + eigen_ai_tk + toegerekend
        per_pers = tot / n_pers if n_pers > 0 else 0
        ws2.append([team_naam, n_pers, eigen_sw, eigen_ai_tk, toegerekend, tot, per_pers])
        for cell in ws2[ws2.max_row]:
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if isinstance(cell.value, float):
                cell.number_format = EURO_HEEL
        grand_eigen_sw += eigen_sw
        grand_eigen_ai += eigen_ai_tk
        grand_toegerekend += toegerekend

    # Onbekend team apart
    onb_df = df_intern[df_intern["Team"] == "Onbekend"]
    if len(onb_df) > 0:
        eigen_sw = float(onb_df[onb_df["Bucket"] == "Software"]["Saldo"].sum()) / n_maanden
        eigen_ai_tk = float(onb_df[onb_df["Bucket"].isin(["AI-tools", "Tokens Intern"])]["Saldo"].sum()) / n_maanden
        tot = eigen_sw + eigen_ai_tk
        ws2.append(["Onbekend / niet toegewezen", "—", eigen_sw, eigen_ai_tk, 0, tot, ""])
        for cell in ws2[ws2.max_row]:
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.fill = GEEL_FILL
            if isinstance(cell.value, float):
                cell.number_format = EURO_HEEL
        grand_eigen_sw += eigen_sw
        grand_eigen_ai += eigen_ai_tk

    # Totaal-rij
    ws2.append(["TOTAAL Intern", totaal_personen, grand_eigen_sw, grand_eigen_ai,
                grand_toegerekend,
                grand_eigen_sw + grand_eigen_ai + grand_toegerekend,
                (grand_eigen_sw + grand_eigen_ai + grand_toegerekend) / totaal_personen if totaal_personen > 0 else 0])
    for cell in ws2[ws2.max_row]:
        cell.font = BOLD_FONT
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
        if isinstance(cell.value, float):
            cell.number_format = EURO_HEEL

    write_blank(ws2)
    # Toelichting onder de tabel
    ws2.append([f"  • Organisatiebrede pool: € {org_sw + org_ai + org_tk:,.0f}/mnd "
                f"= € {org_pp:,.2f}/persoon/mnd (Atlassian, Bitwarden, Loom, etc.)"])
    ws2.cell(row=ws2.max_row, column=1).font = ITALIC_FONT
    for combi_team, sub_teams in GEDEELD_SPLIT.items():
        combi_df = df_intern[df_intern["Team"] == combi_team]
        if len(combi_df) > 0:
            combi_total_pm = float(combi_df["Saldo"].sum()) / n_maanden
            ws2.append([f"  • {combi_team}: € {combi_total_pm:,.0f}/mnd "
                        f"verdeeld over {' en '.join(sub_teams)}"])
            ws2.cell(row=ws2.max_row, column=1).font = ITALIC_FONT

    write_blank(ws2)

    # Sectie E — Projectie
    write_section_title(ws2, "E. Projectie — wat als we groeien? (alleen intern)")
    ws2.append([f"Huidig: {totaal_personen} personen. Per-seat licenties schalen "
                "lineair, tokens schalen ~lineair (variabel — disclaimer)."])
    ws2.cell(row=ws2.max_row, column=1).font = ITALIC_FONT

    ws2.append([" ", f"Huidig ({totaal_personen})",
                f"+1 persoon ({totaal_personen+1})",
                f"+5 personen ({totaal_personen+5})",
                f"+10 personen ({totaal_personen+10})"])
    style_header_row(ws2, ws2.max_row)

    def project_intern(extra_pers):
        nieuw = totaal_personen + extra_pers
        f = nieuw / totaal_personen
        ai_proj = 0.0
        for lev in df_intern[df_intern["Bucket"] == "AI-tools"]["Leverancier"].unique():
            sub = df_intern[df_intern["Leverancier"] == lev]
            cur = float(sub["Saldo"].sum()) / n_maanden
            is_seat = bool(sub["Per_seat"].iloc[0])
            ai_proj += cur * f if is_seat else cur
        sw_proj = 0.0
        for lev in df_intern[df_intern["Bucket"] == "Software"]["Leverancier"].unique():
            sub = df_intern[df_intern["Leverancier"] == lev]
            cur = float(sub["Saldo"].sum()) / n_maanden
            is_seat = bool(sub["Per_seat"].iloc[0])
            sw_proj += cur * f if is_seat else cur
        tk_proj = tokens_intern_total_pm * f  # tokens schalen ~lineair
        return sw_proj, ai_proj, tk_proj

    rijen = [("Software", 0), ("AI-tools", 1), ("Tokens Intern", 2)]
    proj_data = {x: project_intern(x) for x in [0, 1, 5, 10]}
    for label, idx in rijen:
        ws2.append([label, proj_data[0][idx], proj_data[1][idx],
                    proj_data[5][idx], proj_data[10][idx]])
        for cell in ws2[ws2.max_row]:
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if isinstance(cell.value, float):
                cell.number_format = EURO_HEEL
    # Totaal
    totals = [sum(proj_data[x]) for x in [0, 1, 5, 10]]
    ws2.append(["Totaal Intern/mnd"] + totals)
    for cell in ws2[ws2.max_row]:
        cell.font = BOLD_FONT
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
        if isinstance(cell.value, float):
            cell.number_format = EURO_HEEL
    # Per persoon
    pp = [totals[i] / (totaal_personen + x) for i, x in enumerate([0, 1, 5, 10])]
    ws2.append(["Per persoon/mnd"] + pp)
    for cell in ws2[ws2.max_row]:
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        if isinstance(cell.value, float):
            cell.number_format = EURO_HEEL

    auto_width(ws2, max_width=42)
    ws2.freeze_panes = "A3"

    # =====================================================================
    # TAB 3 — Cloud / Tokens Extern
    # =====================================================================
    ws3 = wb.create_sheet("Cloud Extern")
    ws3.sheet_view.showGridLines = False

    ws3.append(["Cloud / Tokens Extern — productiekosten direct voor klant"])
    ws3.cell(row=1, column=1).font = Font(name=ARIAL, bold=True, size=14, color="1F4E79")
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    write_blank(ws3)

    df_extern = df_reg_actief[df_reg_actief["Bucket"] == "Tokens Extern"]

    # KPI extern
    ws3.append([f"Totaal extern YTD: € {extern_ytd:,.0f}   |   Gemiddeld per maand: € {extern_per_maand:,.0f}"])
    ws3.cell(row=ws3.max_row, column=1).font = Font(name=ARIAL, bold=True, size=11, color="1F4E79")
    write_blank(ws3)

    # Per leverancier
    write_section_title(ws3, "Per leverancier")
    ws3.append(["Leverancier", "Gem./maand", "Totaal YTD", "% van extern", "Trend", "Notitie"])
    style_header_row(ws3, ws3.max_row)

    lev_extern = df_extern.groupby("Leverancier")["Saldo"].sum().sort_values(ascending=False)
    for lev_naam, ytd in lev_extern.items():
        sub = df_extern[df_extern["Leverancier"] == lev_naam]
        gem_pm = float(ytd) / n_maanden
        per_period = sub.groupby("Maand_nr")["Saldo"].sum().to_dict()
        trend = trend_indicator(per_period)
        pct = float(ytd) / extern_ytd if extern_ytd else 0
        notitie = sub["Notitie_lev"].iloc[0] or ""
        ws3.append([lev_naam, gem_pm, float(ytd), pct, trend, notitie])
        for cell in ws3[ws3.max_row]:
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.fill = BUCKET_FILL["Tokens Extern"]
        ws3.cell(row=ws3.max_row, column=2).number_format = EURO_HEEL
        ws3.cell(row=ws3.max_row, column=3).number_format = EURO_HEEL
        ws3.cell(row=ws3.max_row, column=4).number_format = PCT_FMT
        ws3.cell(row=ws3.max_row, column=5).alignment = Alignment(horizontal="center")

    ws3.append(["TOTAAL extern", extern_per_maand, extern_ytd, 1.0, "", ""])
    for cell in ws3[ws3.max_row]:
        cell.font = BOLD_FONT
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
    ws3.cell(row=ws3.max_row, column=2).number_format = EURO_HEEL
    ws3.cell(row=ws3.max_row, column=3).number_format = EURO_HEEL
    ws3.cell(row=ws3.max_row, column=4).number_format = PCT_FMT

    write_blank(ws3)

    # Per maand
    write_section_title(ws3, "Per maand")
    header = ["Leverancier"] + [MAAND_NAMEN[m] for m in reguliere_maanden] + ["Totaal YTD"]
    ws3.append(header)
    style_header_row(ws3, ws3.max_row)

    for lev_naam in lev_extern.index:
        sub = df_extern[df_extern["Leverancier"] == lev_naam]
        row = [lev_naam]
        for m in reguliere_maanden:
            row.append(float(sub[sub["Maand_nr"] == m]["Saldo"].sum()))
        row.append(float(sub["Saldo"].sum()))
        ws3.append(row)
        for cell in ws3[ws3.max_row]:
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.fill = BUCKET_FILL["Tokens Extern"]
            if isinstance(cell.value, float):
                cell.number_format = EURO_HEEL

    tot_row = ["TOTAAL"]
    for m in reguliere_maanden:
        tot_row.append(float(df_extern[df_extern["Maand_nr"] == m]["Saldo"].sum()))
    tot_row.append(extern_ytd)
    ws3.append(tot_row)
    for cell in ws3[ws3.max_row]:
        cell.font = BOLD_FONT
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
        if isinstance(cell.value, float):
            cell.number_format = EURO_HEEL

    write_blank(ws3)
    write_blank(ws3)

    # Toelichting
    ws3.append(["Toelichting:"])
    ws3.cell(row=ws3.max_row, column=1).font = BOLD_FONT
    ws3.append(["• Dit zijn variabele kosten die we MAKEN voor de klant — productie/cloud."])
    ws3.append(["• Hoort op grootboekrekening 'Inkoop cloud'."])
    ws3.append(["• Moonlit Legal Technologies = directe leverancier voor het product (~€4k/mnd)."])
    ws3.append(["• Anthropic API = productie-tokens voor klant-features."])
    ws3.append(["• Azure / GCP / AWS / Supabase = hosting/infra van het product."])
    for r in range(ws3.max_row - 4, ws3.max_row + 1):
        ws3.cell(row=r, column=1).font = ITALIC_FONT

    auto_width(ws3, max_width=42)
    ws3.freeze_panes = "A3"

    # =====================================================================
    # TAB 4 — Beslislijst (verplaatsingen)
    # =====================================================================
    ws4 = wb.create_sheet("Beslislijst")
    ws4.sheet_view.showGridLines = False

    ws4.append(["Beslislijst — verplaatsingsoverzicht en open vragen"])
    ws4.cell(row=1, column=1).font = Font(name=ARIAL, bold=True, size=14, color="1F4E79")
    ws4.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    write_blank(ws4)

    # Sectie A: Verplaatsen tussen grootboekrekeningen
    write_section_title(ws4, "A. Verplaatsen — staat hier, hoort hier")

    verkeerd_df = df_reg_actief[df_reg_actief["Verkeerd_geboekt"] == True]

    if len(verkeerd_df) > 0:
        # Eerst overzicht per leverancier
        ws4.append(["Leverancier", "Bucket", "Aantal regels",
                    "Totaal YTD", "Staat in", "Hoort in"])
        style_header_row(ws4, ws4.max_row)

        for lev_naam, sub in verkeerd_df.groupby("Leverancier"):
            bucket = sub["Bucket"].iloc[0]
            staat_in = ", ".join(sorted(sub["Bron"].unique()))
            hoort_in = JUISTE_BRON_VAN_BUCKET.get(bucket, "—")
            ws4.append([lev_naam, bucket, len(sub),
                        float(sub["Saldo"].sum()), staat_in, hoort_in])
            for cell in ws4[ws4.max_row]:
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                cell.fill = ORANJE_FILL
            ws4.cell(row=ws4.max_row, column=4).number_format = EURO_HEEL

        ws4.append(["TOTAAL te verplaatsen", "", len(verkeerd_df),
                    float(verkeerd_df["Saldo"].sum()), "", ""])
        for cell in ws4[ws4.max_row]:
            cell.font = BOLD_FONT
            cell.fill = TOTAL_FILL
            cell.border = THIN_BORDER
        ws4.cell(row=ws4.max_row, column=4).number_format = EURO_HEEL

        write_blank(ws4)

        # Detail per regel
        write_section_title(ws4, "A1. Detail per mutatie")
        ws4.append(["Datum", "Leverancier", "Boeking origineel", "Bedrag",
                    "Staat in", "Hoort in"])
        style_header_row(ws4, ws4.max_row)
        for _, r in verkeerd_df.sort_values(["Leverancier", "Boekstuk"]).iterrows():
            datum = r["Boekstuk"].strftime("%Y-%m-%d") if hasattr(r["Boekstuk"], "strftime") else str(r["Boekstuk"])
            hoort_in = JUISTE_BRON_VAN_BUCKET.get(r["Bucket"], "—")
            ws4.append([datum, r["Leverancier"], str(r["Boeking"]).strip(),
                        float(r["Saldo"]), str(r.get("Bron", "")), hoort_in])
            for cell in ws4[ws4.max_row]:
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                cell.fill = ORANJE_FILL
            ws4.cell(row=ws4.max_row, column=4).number_format = EURO_HEEL
    else:
        ws4.append(["✅ Alles staat op de juiste grootboekrekening — niets te verplaatsen"])
        for cell in ws4[ws4.max_row]:
            cell.font = BOLD_FONT
            cell.fill = GROEN_FILL

    write_blank(ws4)
    write_blank(ws4)

    # Sectie B: Te classificeren
    write_section_title(ws4, "B. Te classificeren — onbekende of onzekere leveranciers")

    onbekend_groep = df_reg_actief[df_reg_actief["Leverancier"] == "Onbekend"]
    # Ook leveranciers met "??" in notitie tonen
    onzeker_levs = [lev["naam"] for lev in leveranciers
                    if "??" in (lev.get("notities") or "")]
    onzeker_df = df_reg_actief[df_reg_actief["Leverancier"].isin(onzeker_levs)]

    if len(onbekend_groep) > 0 or len(onzeker_df) > 0:
        ws4.append(["Leverancier / Boeking", "Totaal YTD", "Bron", "Status / Vraag"])
        style_header_row(ws4, ws4.max_row)

        for boeking, sub in onbekend_groep.groupby("Boeking"):
            tot = float(sub["Saldo"].sum())
            bron_str = ", ".join(sorted(sub["Bron"].unique()))
            ws4.append([str(boeking).strip(), tot, bron_str, "Onbekend — classificeren?"])
            for cell in ws4[ws4.max_row]:
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                cell.fill = GEEL_FILL
            ws4.cell(row=ws4.max_row, column=2).number_format = EURO_HEEL

        for lev_naam in onzeker_levs:
            sub = onzeker_df[onzeker_df["Leverancier"] == lev_naam]
            if len(sub) == 0:
                continue
            tot = float(sub["Saldo"].sum())
            bron_str = ", ".join(sorted(sub["Bron"].unique()))
            notitie = sub["Notitie_lev"].iloc[0]
            ws4.append([lev_naam, tot, bron_str, notitie])
            for cell in ws4[ws4.max_row]:
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                cell.fill = GEEL_FILL
            ws4.cell(row=ws4.max_row, column=2).number_format = EURO_HEEL
    else:
        ws4.append(["✅ Geen openstaande classificatievragen"])
        for cell in ws4[ws4.max_row]:
            cell.font = BOLD_FONT
            cell.fill = GROEN_FILL

    auto_width(ws4, max_width=45)
    ws4.freeze_panes = "A3"

    # =====================================================================
    # TAB 5 — Alle mutaties + legenda
    # =====================================================================
    ws5 = wb.create_sheet("Alle mutaties")

    # Legenda bovenaan
    ws5.append(["LEGENDA"])
    ws5.cell(row=1, column=1).font = Font(name=ARIAL, bold=True, size=12, color="1F4E79")
    ws5.append(["Bucket", "Wat", "Voorbeeld"])
    style_header_row(ws5, ws5.max_row)
    legenda_rijen = [
        ("Software", "Vaste software-abonnementen + niet-software abonnementen", "Atlassian, HubSpot, Cursor, AFAS"),
        ("AI-tools", "AI-diensten met vast abonnement", "DeepL, Lovable, HeyGen, xAI"),
        ("Tokens Intern", "Variabele AI/cloud DOOR HET TEAM (engineering, dienstverbetering)", "Claude Team, ChatGPT Team, Cursor AI usage, Fireflies"),
        ("Tokens Extern", "Variabele cloud/tokens DIRECT VOOR DE KLANT (productie)", "Azure, Google Cloud, Anthropic API, Moonlit, AWS"),
        ("Te classificeren", "Onbekende leverancier — nog te classificeren", "—"),
    ]
    for naam, wat, voorb in legenda_rijen:
        ws5.append([naam, wat, voorb])
        fill = BUCKET_FILL.get(naam)
        for cell in ws5[ws5.max_row]:
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill
    write_blank(ws5)
    ws5.append(["Status:"])
    ws5.cell(row=ws5.max_row, column=1).font = BOLD_FONT
    ws5.append(["actief", "Loopt nog door — telt mee in maandgemiddeldes en alle totalen", ""])
    for cell in ws5[ws5.max_row]:
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
    ws5.append(["stopgezet", "Reeds afgesloten — telt niet mee in actieve totalen, alleen historisch zichtbaar", ""])
    for cell in ws5[ws5.max_row]:
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        cell.fill = STOPGEZET_FILL

    write_blank(ws5)
    ws5.append(["Bron:"])
    ws5.cell(row=ws5.max_row, column=1).font = BOLD_FONT
    ws5.append(["Welke AFAS-export de mutatie afkomstig van is. Doorgaans 'Abonnementen' of 'Inkoop cloud'.", "", ""])
    ws5.cell(row=ws5.max_row, column=1).font = ITALIC_FONT

    write_blank(ws5)
    write_blank(ws5)

    # Hoofd-data
    ws5.append(["Datum", "Maand", "Leverancier", "Bucket", "Status", "Team",
                "Per seat", "Bedrag", "Bron", "Verkeerd geboekt?",
                "Boeking origineel", "Fact.nr.", "Dagboek"])
    style_header_row(ws5, ws5.max_row)
    header_row = ws5.max_row

    for _, r in df.sort_values(["Boekstuk", "Boeking"]).iterrows():
        datum = r["Boekstuk"].strftime("%Y-%m-%d") if hasattr(r["Boekstuk"], "strftime") else str(r["Boekstuk"])
        bucket_label = r["Bucket"] if r["Bucket"] else "Te classificeren"
        verkeerd = "⚠️ Ja" if r.get("Verkeerd_geboekt") else ""
        ws5.append([datum, r["Maand"], r["Leverancier"], bucket_label,
                    r["Status"], r["Team"],
                    "Ja" if r["Per_seat"] else "Nee",
                    float(r["Saldo"]),
                    str(r.get("Bron", "") or ""),
                    verkeerd,
                    str(r["Boeking"]).strip(),
                    str(r.get("Fact.nr.", "") or ""),
                    str(r.get("Dagboek", "") or "")])
        fill = BUCKET_FILL.get(bucket_label)
        for cell in ws5[ws5.max_row]:
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if r["Status"] == "stopgezet":
                cell.fill = STOPGEZET_FILL
            elif fill:
                cell.fill = fill
        ws5.cell(row=ws5.max_row, column=8).number_format = EURO_CENT

    ws5.freeze_panes = ws5.cell(row=header_row + 1, column=1).coordinate
    auto_width(ws5, max_width=40)

    # =====================================================================
    # Opslaan
    # =====================================================================
    wb.save(output_path)

    # Console-samenvatting
    print(f"\n=== Licentie-analyse gegenereerd: {output_path} ===")
    print(f"Maanden in data: {[MAAND_NAMEN[m] for m in reguliere_maanden]}")
    print(f"\nActief YTD:                 € {totaal_actief_ytd:>10,.0f}")
    print(f"  Intern (team)/maand:      € {intern_per_maand:>10,.0f}")
    print(f"    Software/maand:         € {bucket_actief_ytd.get('Software',0)/n_maanden:>10,.0f}")
    print(f"    AI-tools/maand:         € {bucket_actief_ytd.get('AI-tools',0)/n_maanden:>10,.0f}")
    print(f"    Tokens Intern/maand:    € {bucket_actief_ytd.get('Tokens Intern',0)/n_maanden:>10,.0f}")
    print(f"  Extern (klant)/maand:     € {extern_per_maand:>10,.0f}")
    print(f"    Tokens Extern/maand:    € {bucket_actief_ytd.get('Tokens Extern',0)/n_maanden:>10,.0f}")
    print(f"\nStopgezet (historisch YTD): € {totaal_stopgezet_ytd:>10,.0f}")
    if df_corr.shape[0] > 0:
        print(f"Correctieboekingen:         € {float(df_corr['Saldo'].sum()):>10,.0f}")
    if len(verkeerd_df) > 0:
        print(f"\nTe verplaatsen tussen grootboeken: € {float(verkeerd_df['Saldo'].sum()):>10,.0f} ({len(verkeerd_df)} regels)")
    if len(onbekend_groep) > 0:
        print(f"\nOnbekende leveranciers: {onbekend_groep['Boeking'].nunique()}")


if __name__ == "__main__":
    if len(sys.argv) != 5:
        print("Gebruik: python verwerk_licenties.py <input.xlsx> "
              "<leveranciers.json> <teams.json> <output.xlsx>")
        print("Multi-source: gebruik 'pad1=Label1,pad2=Label2' als input.")
        sys.exit(1)
    verwerk(*sys.argv[1:5])
